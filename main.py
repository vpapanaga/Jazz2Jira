import json
import openpyxl

with open('init.conf') as config_file:
    lines=config_file.readlines()
    path_to_excel_file = lines[0].split('=')[-1].strip('\n')
    path_to_folder = lines[1].split('=')[-1].strip('\n')+"workitem."



wb=openpyxl.load_workbook(path_to_excel_file)
# sheet=wb.get_sheet_by_name('Test Plan')
sheet=wb.get_active_sheet()
epic=[]


epic_ids=[ path_to_folder+str(sheet.cell(row=i, column=2).value)+".json"
           for i in range(2,201) if sheet.cell(row=i, column=1).value=='Epic']

epic_list=['dc:description'
            ,'dc:title'
            ,'dc:created'
            ,'rtc_cm:com.ibm.team.workitem.linktype.parentworkitem.children,oslc_cm:label']

story_list=['rtc_cm:com.ibm.team.workitem.linktype.attachment.attachment,oslc_cm:label'
            ,'rtc_cm:com.ibm.team.workitem.linktype.textualReference.textuallyReferenced,oslc_cm:label'
            ,'calm:relatedTestPlan,oslc_cm:label'
            ,'rtc_cm:com.ibm.team.workitem.linktype.parentworkitem.children,oslc_cm:label'
            ,'rtc_cm:com.ibm.team.apt.attribute.acceptance'
            ,'dc:title'
            ,'dc:description']
task_list=['dc:description'
           ,'rtc_cm:timeSpent'
           ,'dc:title']

def get_node_value(file_path,list_node):
    values={}
    try:
        with open(file_path,encoding="utf8") as file:
            values['file_path'] = file_path
            # print("Epic ="+file_path)
            data = json.loads(file.read())
        for node_name in list_node:
            val=node_name.split(',')
            if len(val)>1:
                values[node_name]=[ i['oslc_cm:label'] for i in data['rtc_cm:com.ibm.team.workitem.linktype.parentworkitem.children']]
            else:
                try:
                    values[node_name] = data[val[0]]
                except:
                    values[node_name] =''
    except : return 0
    return values

def get_story_values(list_id,list):
    values={}
    for id in [i.split(':')[0] for i in list_id]:
        rez=get_node_value(file_path=path_to_folder + str(id) + ".json", list_node=list)
        if rez==0:continue
        values[id]=rez
    return values

def get_task_values(list_id):
    value={}
    for k,v in list_id.items():
        # print(k,v['rtc_cm:com.ibm.team.workitem.linktype.parentworkitem.children,oslc_cm:label'])
        if(len(v['rtc_cm:com.ibm.team.workitem.linktype.parentworkitem.children,oslc_cm:label'])==0): continue
        for k,v in get_story_values(v['rtc_cm:com.ibm.team.workitem.linktype.parentworkitem.children,oslc_cm:label'],task_list).items():
            value[k]=v
    return value


# print(task_nodes)
# for key,value in story_nodes.items():
#     print(key+"== "+str(value))


def print_nodes(nodes):
    with open("test.txt", "a",encoding="utf8") as file:
        for key,value in nodes.items():
            if key=='rtc_cm:com.ibm.team.workitem.linktype.parentworkitem.children,oslc_cm:label':
                file.write("\n\t\t\t" + str(key)+'\n')
                for i in value:
                    id_story=str(i).split(':')[0]
                    file.write("\t"*3+str(i)+'\n')
                    if(story_nodes.get(id_story,0)==0):
                        print(id_story)
                    for k,v in story_nodes.get(id_story).items():
                        if k == 'rtc_cm:com.ibm.team.workitem.linktype.parentworkitem.children,oslc_cm:label' and len(v)>0:
                            file.write("\t"*6 + str(k)+'\n')
                            for j in v:
                                id_task=j.split(':')[0]
                                # print(id_task)
                                # print(j)
                                try:
                                    if(0== task_nodes.get(id_task,0)): continue
                                    for task_key,task_value in task_nodes.get(id_task).items():
                                        file.write("\t"*8+task_key+"--- "+str(task_value)+'\n')
                                except:
                                    print('sssss')
                        file.write("\t"*6 + str(k)+' '+str(v)+'\n')

            else:
                file.write(key+"== "+str(value)+'\n')
        file.write('-'*100+'\n')

# for ep in epic_ids:
#     print(ep)
# print(epic_ids[16])
for epcn in epic_ids:
    print(epcn)
    epic_nodes=get_node_value(file_path=epcn,list_node=epic_list)
    story_nodes=get_story_values(epic_nodes['rtc_cm:com.ibm.team.workitem.linktype.parentworkitem.children,oslc_cm:label'],story_list)
    task_nodes=get_task_values(story_nodes)

    # for k,v in story_nodes.items():
    #     for j in v.get('rtc_cm:com.ibm.team.workitem.linktype.parentworkitem.children,oslc_cm:label'):
    #         print(j)

# print(epic_nodes)
#     for key, value in epic_nodes.items():
#         # print(key)
#         if key == 'rtc_cm:com.ibm.team.workitem.linktype.parentworkitem.children,oslc_cm:label':
#             for i in value:
#                 print(epic_nodes.get('file_path'),i)



    # print(story_nodes)
    # print(task_nodes)

    print_nodes(epic_nodes)

# for key,value in story_nodes.items():
#     print(key+"== "+str(value))

# print(epic_ids)


# for path in epic_ids:
#     with open(path) as file:
#         print("Epic ="+path)
#         data = json.loads(file.read())
#         with open("test.txt","a") as file:
#             file.write("\n\tEPIC\n\tfile path ="+path+"\n\tuser_creator = "+data['dc:creator']['rdf:resource']+'\n')
#             file.write("\n\t\tSTORY ->")
#             story_ids=[str(i['rdf:resource'].split('/')[-1]) for i in data['rtc_cm:com.ibm.team.workitem.linktype.parentworkitem.children']]
#             for story_id in story_ids:
#                 file.write("\n\t\t\t file path ="+path_to_folder + story_id+".json")
#                 file.write("\n\t\t\t -"+story_id)
#                 print("Story=="+path_to_folder+str(story_id)+".json")
#                 with open(path_to_folder+str(story_id)+".json",encoding="utf8") as task_file:
#                     data_task = json.loads(task_file.read())
#                     dec = {str(i['rdf:resource'].split('/')[-1]) : i["oslc_cm:label"] for i in
#                                data_task['rtc_cm:com.ibm.team.workitem.linktype.parentworkitem.children']}
#                     file.write("\n\t\t\t\t TASK ->" )
#                     for key,value in dec.items():
#                         file.write("\n\t\t\t\t\t file path =" +path_to_folder+str(key)+".json")
#                         file.write("\n\t\t\t\t\t id ->"+key)
#                         file.write("\n\t\t\t\t\t value ->"+ value)
#                     # print(data_task)



